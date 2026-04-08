from __future__ import annotations

import io

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from services.models import ComparisonResult, ForecastData
from utils.constants import EXPORT_SHEET_NAMES


class ExportService:
    @staticmethod
    def build_excel_export(forecasts: list[ForecastData], comparison: ComparisonResult) -> bytes:
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            comparison.summary.to_excel(writer, sheet_name=EXPORT_SHEET_NAMES["summary"], index=False)

            for index, forecast in enumerate(forecasts, start=1):
                forecast.consolidated_data.to_excel(writer, sheet_name=f"Previsao_{index}", index=False)

            comparison.detail.to_excel(writer, sheet_name=EXPORT_SHEET_NAMES["detail"], index=False)

            all_unclassified = pd.concat([forecast.unclassified_data.assign(previsao=forecast.label) for forecast in forecasts], ignore_index=True)
            if not all_unclassified.empty:
                all_unclassified.to_excel(writer, sheet_name=EXPORT_SHEET_NAMES["unclassified"], index=False)

            ExportService._format_workbook(writer)
        return output.getvalue()

    @staticmethod
    def _format_workbook(writer: pd.ExcelWriter) -> None:
        workbook = writer.book
        header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="E5E7EB")
        border = Border(bottom=thin)

        for ws in workbook.worksheets:
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            for col_idx, column_cells in enumerate(ws.columns, start=1):
                max_length = 0
                for cell in column_cells:
                    value = "" if cell.value is None else str(cell.value)
                    max_length = max(max_length, len(value))
                    if col_idx > 1:
                        cell.alignment = Alignment(vertical="top")
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_length + 2, 14), 42)
